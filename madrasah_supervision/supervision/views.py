from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count, Avg, Q, Sum
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.contrib import messages
from django.template.loader import get_template
from io import BytesIO
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.utils import PlotlyJSONEncoder
import json
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.pdfgen import canvas

from .models import *
from .forms import *

def is_supervisor(user):
    return hasattr(user, 'supervisor')

def is_kepala_madrasah(user):
    return user.groups.filter(name='Kepala Madrasah').exists() or hasattr(user, 'madrasah_kepala')


# ============= DASHBOARD =============

@login_required
def dashboard(request):
    context = {}
    
    if request.user.is_staff:
        context.update({
            'total_madrasah': Madrasah.objects.count(),
            'total_guru': Guru.objects.count(),
            'total_supervisi': HasilSupervisi.objects.count(),
            'total_supervisor': Supervisor.objects.count(),
        })
    elif hasattr(request.user, 'supervisor'):
        supervisor = request.user.supervisor
        madrasah_binaan = supervisor.madrasah_binaan.all()
        context.update({
            'total_madrasah': madrasah_binaan.count(),
            'total_guru': Guru.objects.filter(madrasah__in=madrasah_binaan).count(),
            'total_supervisi': HasilSupervisi.objects.filter(supervisor=supervisor).count(),
            'rata_rata_nilai': HasilSupervisi.objects.filter(supervisor=supervisor).aggregate(Avg('nilai_total'))['nilai_total__avg'] or 0,
            'supervisor': supervisor,
        })
    elif is_kepala_madrasah(request.user):
        try:
            madrasah = Madrasah.objects.get(kepala_madrasah__icontains=request.user.get_full_name())
            context.update({
                'madrasah': madrasah,
                'total_tagihan': TagihanPengawas.objects.filter(madrasah=madrasah).count(),
                'tagihan_pending': TagihanPengawas.objects.filter(madrasah=madrasah, status='pending').count(),
                'tagihan_terbaru': TagihanPengawas.objects.filter(madrasah=madrasah)[:5],
            })
        except Madrasah.DoesNotExist:
            pass
    
    return render(request, 'supervision/dashboard.html', context)


# ============= SUPERVISI AKADEMIK =============

@login_required
@user_passes_test(is_supervisor)
def input_supervisi(request):
    supervisor = request.user.supervisor
    
    # Cek apakah ini edit (ada parameter id)
    supervisi_id = request.GET.get('id') or request.POST.get('id')
    supervisi = None
    
    if supervisi_id:
        supervisi = get_object_or_404(HasilSupervisi, id=supervisi_id, supervisor=supervisor)
    
    if request.method == 'POST':
        madrasah_id = request.POST.get('madrasah')
        guru_id = request.POST.get('guru')
        tanggal = request.POST.get('tanggal')
        semester = request.POST.get('semester')
        tahun_ajaran = request.POST.get('tahun_ajaran')
        rekomendasi = request.POST.get('rekomendasi')
        tindak_lanjut = request.POST.get('tindak_lanjut')
        file_pendukung = request.FILES.get('file_pendukung')
        save_komponen = request.POST.get('save_komponen')
        
        # Kumpulkan nilai yang diisi saja (abaikan yang kosong)
        nilai_per_indikator = {}
        for key, value in request.POST.items():
            if key.startswith('nilai_') and value.strip():
                try:
                    nilai = float(value)
                    if 0 <= nilai <= 100:
                        nilai_per_indikator[key.replace('nilai_', '')] = nilai
                except ValueError:
                    pass
        
        if supervisi:
            # Update data yang sudah ada
            supervisi.madrasah_id = madrasah_id
            supervisi.guru_id = guru_id
            supervisi.tanggal = tanggal
            supervisi.semester = semester
            supervisi.tahun_ajaran = tahun_ajaran
            if rekomendasi:
                supervisi.rekomendasi = rekomendasi
            if tindak_lanjut:
                supervisi.tindak_lanjut = tindak_lanjut
            if file_pendukung:
                supervisi.file_pendukung = file_pendukung
            # Merge nilai lama dengan nilai baru
            existing_nilai = supervisi.nilai_per_indikator or {}
            existing_nilai.update(nilai_per_indikator)
            supervisi.nilai_per_indikator = existing_nilai
            supervisi.save()
            
            if save_komponen:
                # Jika simpan per komponen, kembalikan ke halaman yang sama
                messages.success(request, f'Nilai komponen berhasil disimpan!')
                return redirect(f'/input-supervisi/?id={supervisi.id}')
            else:
                messages.success(request, 'Data supervisi berhasil diperbarui!')
        else:
            # Buat baru
            supervisi = HasilSupervisi.objects.create(
                supervisor=supervisor,
                madrasah_id=madrasah_id,
                guru_id=guru_id,
                tanggal=tanggal,
                semester=semester,
                tahun_ajaran=tahun_ajaran,
                nilai_per_indikator=nilai_per_indikator,
                rekomendasi=rekomendasi,
                tindak_lanjut=tindak_lanjut,
                file_pendukung=file_pendukung
            )
            messages.success(request, 'Data supervisi berhasil disimpan!')
        
        return redirect('/riwayat-supervisi/')
    
    # GET request - tampilkan form
    form = HasilSupervisiForm()
    form.fields['madrasah'].queryset = supervisor.madrasah_binaan.all()
    
    instrumen = InstrumenSupervisi.objects.all().order_by('komponen')
    
    # Jika edit, isi nilai yang sudah ada
    nilai_existing = {}
    if supervisi:
        nilai_existing = supervisi.nilai_per_indikator or {}
        # Pre-select madrasah dan guru
        form.fields['madrasah'].initial = supervisi.madrasah.id
        form.fields['guru'].queryset = Guru.objects.filter(madrasah=supervisi.madrasah)
        form.fields['guru'].initial = supervisi.guru.id
    
    return render(request, 'supervision/input_supervisi.html', {
        'form': form,
        'instrumen': instrumen,
        'supervisi': supervisi,
        'nilai_existing': nilai_existing,
    })

@login_required
@user_passes_test(is_supervisor)
def riwayat_supervisi(request):
    supervisor = request.user.supervisor
    supervisi_list = HasilSupervisi.objects.filter(supervisor=supervisor).order_by('-tanggal')
    
    paginator = Paginator(supervisi_list, 20)
    page_number = request.GET.get('page')
    supervisi = paginator.get_page(page_number)
    
    return render(request, 'supervision/riwayat_supervisi.html', {'supervisi': supervisi})


@login_required
def detail_supervisi(request, pk):
    supervisi = get_object_or_404(HasilSupervisi, pk=pk)
    
    if hasattr(request.user, 'supervisor') and supervisi.supervisor != request.user.supervisor:
        if not request.user.is_staff:
            messages.error(request, 'Anda tidak memiliki akses ke data ini')
            return redirect('riwayat_supervisi')
    
    return render(request, 'supervision/detail_supervisi.html', {'supervisi': supervisi})


@login_required
@user_passes_test(is_supervisor)
def daftar_madrasah_binaan(request):
    supervisor = request.user.supervisor
    madrasah_list = supervisor.madrasah_binaan.all()
    
    data_madrasah = []
    for m in madrasah_list:
        data_madrasah.append({
            'madrasah': m,
            'jumlah_guru': Guru.objects.filter(madrasah=m).count(),
            'jumlah_supervisi': HasilSupervisi.objects.filter(madrasah=m, supervisor=supervisor).count(),
            'rata_nilai': HasilSupervisi.objects.filter(madrasah=m, supervisor=supervisor).aggregate(Avg('nilai_total'))['nilai_total__avg'] or 0,
        })
    
    return render(request, 'supervision/daftar_madrasah.html', {'data_madrasah': data_madrasah})


@login_required
@user_passes_test(is_supervisor)
def daftar_guru_binaan(request, madrasah_id=None):
    supervisor = request.user.supervisor
    
    if madrasah_id:
        madrasah = get_object_or_404(Madrasah, id=madrasah_id)
        guru_list = Guru.objects.filter(madrasah=madrasah)
    else:
        guru_list = Guru.objects.filter(madrasah__in=supervisor.madrasah_binaan.all())
    
    data_guru = []
    for g in guru_list:
        supervisi_terakhir = HasilSupervisi.objects.filter(guru=g, supervisor=supervisor).order_by('-tanggal').first()
        data_guru.append({
            'guru': g,
            'jumlah_supervisi': HasilSupervisi.objects.filter(guru=g, supervisor=supervisor).count(),
            'nilai_terakhir': supervisi_terakhir.nilai_total if supervisi_terakhir else '-',
            'predikat_terakhir': supervisi_terakhir.predikat if supervisi_terakhir else '-',
        })
    
    return render(request, 'supervision/daftar_guru.html', {'data_guru': data_guru})


@login_required
@user_passes_test(is_supervisor)
def generate_laporan_pdf(request, madrasah_id, guru_id=None):
    madrasah = get_object_or_404(Madrasah, id=madrasah_id)
    supervisor = request.user.supervisor
    
    response = HttpResponse(content_type='text/html')
    response['Content-Disposition'] = f'inline; filename="Laporan_Supervisi_{madrasah.npsn}_{datetime.now().strftime("%Y%m%d")}.html"'
    
    if guru_id:
        guru = get_object_or_404(Guru, id=guru_id)
        hasil_supervisi = HasilSupervisi.objects.filter(supervisor=supervisor, madrasah=madrasah, guru=guru).order_by('-tanggal')
        title = f"Laporan Supervisi {guru.nama}"
    else:
        hasil_supervisi = HasilSupervisi.objects.filter(supervisor=supervisor, madrasah=madrasah).order_by('-tanggal')
        title = f"Laporan Supervisi {madrasah.nama}"
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head><title>{title}</title><style>body{{font-family:Arial;margin:40px;}}h1{{text-align:center;}}</style></head>
    <body><h1>{title}</h1><p>Madrasah: {madrasah.nama}</p><p>Tanggal Cetak: {datetime.now().strftime('%d %B %Y')}</p>
    """
    for hs in hasil_supervisi:
        html += f"<h2>Supervisi: {hs.tanggal}</h2><p>Guru: {hs.guru.nama}</p><p>Nilai: {hs.nilai_total}</p><p>Predikat: {hs.predikat}</p><p>Rekomendasi: {hs.rekomendasi}</p>"
    html += "</body></html>"
    response.write(html)
    return response


@staff_member_required
def import_data(request):
    if request.method == 'POST':
        form = ImportDataForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                df = pd.read_excel(request.FILES['file'])
                messages.success(request, f'Data berhasil diimport! {len(df)} record diproses.')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
            return redirect('dashboard')
    else:
        form = ImportDataForm()
    return render(request, 'supervision/import_data.html', {'form': form})


@staff_member_required
def export_data(request, model_name):
    if model_name == 'madrasah':
        queryset = Madrasah.objects.all()
        data = list(queryset.values('npsn', 'nama', 'jenis', 'status', 'alamat', 'kecamatan', 'kabupaten', 'kepala_madrasah', 'akreditasi'))
        filename = f'madrasah_{datetime.now().strftime("%Y%m%d")}.xlsx'
    elif model_name == 'guru':
        queryset = Guru.objects.all()
        data = list(queryset.values('NUPTK', 'nama', 'madrasah__nama', 'jenis_kelamin', 'bidang_studi'))
        filename = f'guru_{datetime.now().strftime("%Y%m%d")}.xlsx'
    elif model_name == 'supervisi':
        queryset = HasilSupervisi.objects.all()
        data = list(queryset.values('guru__nama', 'madrasah__nama', 'tanggal', 'nilai_total', 'predikat'))
        filename = f'supervisi_{datetime.now().strftime("%Y%m%d")}.xlsx'
    else:
        return redirect('dashboard')
    
    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    df.to_excel(response, index=False)
    return response


@login_required
def kelola_tagihan(request):
    if is_supervisor(request.user):
        supervisor = request.user.supervisor
        
        if request.method == 'POST':
            nomor_tagihan = request.POST.get('nomor_tagihan')
            madrasah_id = request.POST.get('madrasah')
            jatuh_tempo = request.POST.get('jatuh_tempo')
            deskripsi = request.POST.get('deskripsi')
            jumlah = request.POST.get('jumlah')
            file_tagihan = request.FILES.get('file_tagihan')
            
            TagihanPengawas.objects.create(
                madrasah_id=madrasah_id,
                supervisor=supervisor,
                nomor_tagihan=nomor_tagihan,
                jatuh_tempo=jatuh_tempo,
                deskripsi=deskripsi,
                jumlah=jumlah,
                file_tagihan=file_tagihan
            )
            messages.success(request, 'Tagihan berhasil dibuat!')
            return redirect('/kelola-tagihan/')
        
        tagihan = TagihanPengawas.objects.filter(supervisor=supervisor).order_by('-tanggal')
        return render(request, 'supervision/kelola_tagihan.html', {
            'tagihan': tagihan,
            'madrasah_list': supervisor.madrasah_binaan.all(),
            'is_supervisor': True,
        })
    
    elif is_kepala_madrasah(request.user):
        try:
            madrasah = Madrasah.objects.get(kepala_madrasah__icontains=request.user.get_full_name())
            
            if request.method == 'POST' and 'upload_bukti' in request.FILES:
                tagihan_id = request.POST.get('tagihan_id')
                tagihan_obj = get_object_or_404(TagihanPengawas, id=tagihan_id, madrasah=madrasah)
                tagihan_obj.bukti_bayar = request.FILES['upload_bukti']
                tagihan_obj.status = 'dibayar'
                tagihan_obj.save()
                messages.success(request, 'Bukti pembayaran berhasil diupload!')
                return redirect('/kelola-tagihan/')
            
            tagihan = TagihanPengawas.objects.filter(madrasah=madrasah).order_by('-tanggal')
            return render(request, 'supervision/kelola_tagihan.html', {
                'tagihan': tagihan,
                'madrasah': madrasah,
                'is_supervisor': False,
            })
        except Madrasah.DoesNotExist:
            messages.warning(request, 'Data madrasah tidak ditemukan')
            return redirect('/')
    
    return redirect('/')

# ============= DATA BINAAN =============

@login_required
@user_passes_test(is_supervisor)
def daftar_madrasah_binaan(request):
    supervisor = request.user.supervisor
    madrasah_list = supervisor.madrasah_binaan.all()
    
    data_madrasah = []
    for m in madrasah_list:
        data_madrasah.append({
            'madrasah': m,
            'jumlah_guru': Guru.objects.filter(madrasah=m).count(),
            'jumlah_supervisi': HasilSupervisi.objects.filter(madrasah=m, supervisor=supervisor).count(),
            'rata_nilai': HasilSupervisi.objects.filter(madrasah=m, supervisor=supervisor).aggregate(Avg('nilai_total'))['nilai_total__avg'] or 0,
        })
    
    return render(request, 'supervision/daftar_madrasah.html', {'data_madrasah': data_madrasah})


@login_required
@user_passes_test(is_supervisor)
def daftar_guru_binaan(request, madrasah_id=None):
    supervisor = request.user.supervisor
    
    if madrasah_id:
        madrasah = get_object_or_404(Madrasah, id=madrasah_id)
        guru_list = Guru.objects.filter(madrasah=madrasah)
    else:
        guru_list = Guru.objects.filter(madrasah__in=supervisor.madrasah_binaan.all())
    
    data_guru = []
    for g in guru_list:
        supervisi_terakhir = HasilSupervisi.objects.filter(guru=g, supervisor=supervisor).order_by('-tanggal').first()
        data_guru.append({
            'guru': g,
            'jumlah_supervisi': HasilSupervisi.objects.filter(guru=g, supervisor=supervisor).count(),
            'nilai_terakhir': supervisi_terakhir.nilai_total if supervisi_terakhir else '-',
            'predikat_terakhir': supervisi_terakhir.predikat if supervisi_terakhir else '-',
        })
    
    return render(request, 'supervision/daftar_guru.html', {'data_guru': data_guru})


# ============= LAPORAN PDF =============

@login_required
@user_passes_test(is_supervisor)
def generate_laporan_pdf(request, madrasah_id, guru_id=None):
    madrasah = get_object_or_404(Madrasah, id=madrasah_id)
    supervisor = request.user.supervisor
    
    response = HttpResponse(content_type='text/html')
    response['Content-Disposition'] = f'inline; filename="Laporan_Supervisi_{madrasah.npsn}_{datetime.now().strftime("%Y%m%d")}.html"'
    
    if guru_id:
        guru = get_object_or_404(Guru, id=guru_id)
        hasil_supervisi = HasilSupervisi.objects.filter(supervisor=supervisor, madrasah=madrasah, guru=guru).order_by('-tanggal')
        title = f"Laporan Supervisi {guru.nama}"
    else:
        hasil_supervisi = HasilSupervisi.objects.filter(supervisor=supervisor, madrasah=madrasah).order_by('-tanggal')
        title = f"Laporan Supervisi {madrasah.nama}"
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>{title}</title>
        <style>
            body {{ font-family: 'Times New Roman', Arial, sans-serif; margin: 40px; }}
            h1 {{ text-align: center; color: #2c3e50; }}
            h2 {{ color: #34495e; border-bottom: 1px solid #3498db; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; }}
            th {{ background-color: #3498db; color: white; }}
            .footer {{ text-align: center; margin-top: 50px; font-size: 12px; }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <p>Madrasah: {madrasah.nama} (NPSN: {madrasah.npsn})</p>
        <p>Tanggal Cetak: {datetime.now().strftime('%d %B %Y')}</p>
    """
    
    for hs in hasil_supervisi:
        html += f"""
        <h2>Supervisi Tanggal: {hs.tanggal.strftime('%d %B %Y')}</h2>
        <p>Guru: {hs.guru.nama}</p>
        <p>Nilai Total: <strong>{hs.nilai_total}</strong></p>
        <p>Predikat: {hs.predikat}</p>
        <h3>Rekomendasi:</h3>
        <p>{hs.rekomendasi}</p>
        """
    
    html += f"""
        <div class="footer">
            Laporan ini dihasilkan dari Sistem Pengawasan Madrasah
        </div>
    </body>
    </html>
    """
    
    response.write(html)
    return response


# ============= IMPORT/EXPORT =============

@staff_member_required
def import_data(request):
    if request.method == 'POST':
        form = ImportDataForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                df = pd.read_excel(request.FILES['file'])
                for _, row in df.iterrows():
                    madrasah, _ = Madrasah.objects.get_or_create(
                        npsn=str(row['npsn']),
                        defaults={
                            'nama': row['nama_madrasah'],
                            'jenis': row['jenis'],
                            'status': row['status'],
                            'alamat': row['alamat'],
                            'desa': row['desa'],
                            'kecamatan': row['kecamatan'],
                            'kabupaten': row['kabupaten'],
                            'provinsi': row['provinsi'],
                            'kode_pos': str(row['kode_pos']),
                            'kepala_madrasah': row['kepala_madrasah'],
                            'nip_kepala': str(row['nip_kepala']),
                            'akreditasi': row['akreditasi'],
                        }
                    )
                    Guru.objects.get_or_create(
                        NUPTK=str(row['nuptk']),
                        defaults={
                            'nama': row['nama_guru'],
                            'madrasah': madrasah,
                            'jenis_kelamin': row['jenis_kelamin'],
                            'tempat_lahir': row['tempat_lahir'],
                            'tanggal_lahir': datetime.strptime(str(row['tanggal_lahir']), '%Y-%m-%d').date(),
                            'pendidikan_terakhir': row['pendidikan_terakhir'],
                            'bidang_studi': row['bidang_studi'],
                        }
                    )
                messages.success(request, 'Data berhasil diimport!')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
            return redirect('/')
    else:
        form = ImportDataForm()
    
    return render(request, 'supervision/import_data.html', {'form': form})

def get_guru_by_madrasah(request):
    madrasah_id = request.GET.get('madrasah_id')
    if madrasah_id:
        guru_list = Guru.objects.filter(madrasah_id=madrasah_id).values('id', 'nama', 'NUPTK')
        return JsonResponse(list(guru_list), safe=False)
    return JsonResponse([], safe=False)

@staff_member_required
def export_data(request, model_name):
    if model_name == 'madrasah':
        queryset = Madrasah.objects.all()
        data = list(queryset.values('npsn', 'nama', 'jenis', 'status', 'alamat', 'kecamatan', 'kabupaten', 'kepala_madrasah', 'akreditasi'))
        filename = f'madrasah_{datetime.now().strftime("%Y%m%d")}.xlsx'
    elif model_name == 'guru':
        queryset = Guru.objects.all()
        data = list(queryset.values('NUPTK', 'nama', 'madrasah__nama', 'jenis_kelamin', 'bidang_studi'))
        filename = f'guru_{datetime.now().strftime("%Y%m%d")}.xlsx'
    elif model_name == 'supervisi':
        queryset = HasilSupervisi.objects.all()
        data = list(queryset.values('guru__nama', 'madrasah__nama', 'tanggal', 'nilai_total', 'predikat'))
        filename = f'supervisi_{datetime.now().strftime("%Y%m%d")}.xlsx'
    else:
        return redirect('/')
    
    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    df.to_excel(response, index=False)
    return response


# ============= SUPERVISI MANAJERIAL =============

@login_required
@user_passes_test(is_supervisor)
def input_supervisi_manajerial(request):
    supervisor = request.user.supervisor
    
    # Cek apakah ini edit (ada parameter id)
    supervisi_id = request.GET.get('id') or request.POST.get('id')
    supervisi = None
    
    if supervisi_id:
        supervisi = get_object_or_404(HasilSupervisiManajerial, id=supervisi_id, supervisor=supervisor)
    
    if request.method == 'POST':
        madrasah_id = request.POST.get('madrasah')
        tanggal = request.POST.get('tanggal')
        semester = request.POST.get('semester')
        tahun_ajaran = request.POST.get('tahun_ajaran')
        rekomendasi = request.POST.get('rekomendasi')
        tindak_lanjut = request.POST.get('tindak_lanjut')
        file_pendukung = request.FILES.get('file_pendukung')
        save_komponen = request.POST.get('save_komponen')
        
        # Kumpulkan nilai yang diisi saja (abaikan yang kosong)
        nilai_per_indikator = {}
        for key, value in request.POST.items():
            if key.startswith('nilai_') and value.strip():
                try:
                    nilai = float(value)
                    if 0 <= nilai <= 100:
                        nilai_per_indikator[key.replace('nilai_', '')] = nilai
                except ValueError:
                    pass
        
        if supervisi:
            # Update data yang sudah ada
            supervisi.madrasah_id = madrasah_id
            supervisi.tanggal = tanggal
            supervisi.semester = semester
            supervisi.tahun_ajaran = tahun_ajaran
            if rekomendasi:
                supervisi.rekomendasi = rekomendasi
            if tindak_lanjut:
                supervisi.tindak_lanjut = tindak_lanjut
            if file_pendukung:
                supervisi.file_pendukung = file_pendukung
            # Merge nilai lama dengan nilai baru
            existing_nilai = supervisi.nilai_per_indikator or {}
            existing_nilai.update(nilai_per_indikator)
            supervisi.nilai_per_indikator = existing_nilai
            supervisi.save()
            
            if save_komponen:
                # Jika simpan per komponen, kembalikan ke halaman yang sama
                messages.success(request, f'Nilai komponen berhasil disimpan!')
                return redirect(f'/input-supervisi-manajerial/?id={supervisi.id}')
            else:
                messages.success(request, 'Data supervisi manajerial berhasil diperbarui!')
        else:
            # Buat baru
            supervisi = HasilSupervisiManajerial.objects.create(
                supervisor=supervisor,
                madrasah_id=madrasah_id,
                tanggal=tanggal,
                semester=semester,
                tahun_ajaran=tahun_ajaran,
                nilai_per_indikator=nilai_per_indikator,
                rekomendasi=rekomendasi,
                tindak_lanjut=tindak_lanjut,
                file_pendukung=file_pendukung
            )
            messages.success(request, 'Data supervisi manajerial berhasil disimpan!')
        
        return redirect('/riwayat-supervisi-manajerial/')
    
    # GET request - tampilkan form
    instrumen = InstrumenManajerial.objects.all().order_by('komponen')
    madrasah_list = supervisor.madrasah_binaan.all()
    
    # Jika edit, isi nilai yang sudah ada
    nilai_existing = {}
    if supervisi:
        nilai_existing = supervisi.nilai_per_indikator or {}
    
    return render(request, 'supervision/input_supervisi_manajerial.html', {
        'madrasah_list': madrasah_list,
        'instrumen': instrumen,
        'supervisi': supervisi,
        'nilai_existing': nilai_existing,
    })

@login_required
@user_passes_test(is_supervisor)
def riwayat_supervisi_manajerial(request):
    supervisor = request.user.supervisor
    supervisi_list = HasilSupervisiManajerial.objects.filter(supervisor=supervisor).order_by('-tanggal')
    
    paginator = Paginator(supervisi_list, 20)
    page_number = request.GET.get('page')
    supervisi = paginator.get_page(page_number)
    
    return render(request, 'supervision/riwayat_supervisi_manajerial.html', {'supervisi': supervisi})


@login_required
@user_passes_test(is_supervisor)
def detail_supervisi_manajerial(request, pk):
    supervisi = get_object_or_404(HasilSupervisiManajerial, pk=pk)
    instrumen_list = InstrumenManajerial.objects.all()
    
    if supervisi.supervisor != request.user.supervisor and not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses ke data ini')
        return redirect('riwayat_supervisi_manajerial')
    
    return render(request, 'supervision/detail_supervisi_manajerial.html', {
        'supervisi': supervisi,
        'instrumen_list': instrumen_list,
    })


@login_required
@user_passes_test(is_supervisor)
def laporan_manajerial_pdf(request, pk):
    supervisi = get_object_or_404(HasilSupervisiManajerial, pk=pk)
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Laporan Supervisi Manajerial</title>
        <style>
            body {{ font-family: 'Times New Roman', Arial, sans-serif; margin: 40px; }}
            h1 {{ text-align: center; color: #28a745; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; }}
            th {{ background-color: #28a745; color: white; }}
        </style>
    </head>
    <body>
        <h1>LAPORAN SUPERVISI MANAJERIAL</h1>
        <p>Madrasah: {supervisi.madrasah.nama}</p>
        <p>Tanggal: {supervisi.tanggal.strftime('%d %B %Y')}</p>
        <p>Semester: {supervisi.semester} {supervisi.tahun_ajaran}</p>
        <p>Nilai Total: <strong>{supervisi.nilai_total}</strong></p>
        <p>Predikat: {supervisi.predikat}</p>
        <h3>Rekomendasi</h3>
        <p>{supervisi.rekomendasi}</p>
        <h3>Tindak Lanjut</h3>
        <p>{supervisi.tindak_lanjut or '-'}</p>
        <div class="footer" style="margin-top:50px; text-align:center;">
            Laporan dihasilkan dari Sistem Pengawasan Madrasah<br>
            {datetime.now().strftime('%d %B %Y')}
        </div>
    </body>
    </html>
    """
    
    response = HttpResponse(html, content_type='text/html')
    response['Content-Disposition'] = f'inline; filename="Laporan_Manajerial_{supervisi.madrasah.npsn}_{supervisi.tanggal.strftime("%Y%m%d")}.html"'
    return response


@login_required
def dashboard_snp(request):
    """Dashboard Pemantauan 8 Standar Nasional Pendidikan"""
    if request.user.is_staff:
        madrasah_list = Madrasah.objects.all()
    elif is_supervisor(request.user):
        madrasah_list = request.user.supervisor.madrasah_binaan.all()
    else:
        madrasah_list = []
    
    data_snp = []
    for madrasah in madrasah_list:
        pemantauan = {}
        for standar_code, standar_name in PemantauanSNP.STANDAR_CHOICES:
            try:
                data = PemantauanSNP.objects.filter(
                    madrasah=madrasah, 
                    standar=standar_code
                ).order_by('-updated_at').first()
                pemantauan[standar_code] = {
                    'status': data.status if data else 'belum',
                    'skor': data.skor if data else 0
                }
            except:
                pemantauan[standar_code] = {'status': 'belum', 'skor': 0}
        
        data_snp.append({
            'madrasah': madrasah,
            'pemantauan': pemantauan
        })
    
    return render(request, 'supervision/dashboard_snp.html', {'data_snp': data_snp})
# ============= 8 SNP MADRASAH =============

@login_required
def dashboard_snp(request):
    if request.user.is_staff:
        madrasah_list = Madrasah.objects.all()
    elif is_supervisor(request.user):
        madrasah_list = request.user.supervisor.madrasah_binaan.all()
    else:
        madrasah_list = Madrasah.objects.filter(kepala_madrasah__icontains=request.user.get_full_name())
    
    data_snp = []
    for madrasah in madrasah_list:
        pemantauan = {}
        for standar_code, standar_name in PemantauanSNP.STANDAR_CHOICES:
            try:
                data = PemantauanSNP.objects.filter(
                    madrasah=madrasah, standar=standar_code
                ).order_by('-updated_at').first()
                pemantauan[standar_code] = {'status': data.status, 'skor': data.skor} if data else {'status': 'belum', 'skor': 0}
            except:
                pemantauan[standar_code] = {'status': 'belum', 'skor': 0}
        data_snp.append({'madrasah': madrasah, 'pemantauan': pemantauan})
    
    return render(request, 'supervision/dashboard_snp.html', {'data_snp': data_snp})


# ============= KELOLA TAGIHAN =============

@login_required
def kelola_tagihan(request):
    if is_supervisor(request.user):
        supervisor = request.user.supervisor
        
        if request.method == 'POST':
            nomor_tagihan = request.POST.get('nomor_tagihan')
            madrasah_id = request.POST.get('madrasah')
            jatuh_tempo = request.POST.get('jatuh_tempo')
            deskripsi = request.POST.get('deskripsi')
            jumlah = request.POST.get('jumlah')
            file_tagihan = request.FILES.get('file_tagihan')
            
            TagihanPengawas.objects.create(
                madrasah_id=madrasah_id,
                supervisor=supervisor,
                nomor_tagihan=nomor_tagihan,
                jatuh_tempo=jatuh_tempo,
                deskripsi=deskripsi,
                jumlah=jumlah,
                file_tagihan=file_tagihan
            )
            messages.success(request, 'Tagihan berhasil dibuat!')
            return redirect('/kelola-tagihan/')
        
        tagihan = TagihanPengawas.objects.filter(supervisor=supervisor).order_by('-tanggal')
        return render(request, 'supervision/kelola_tagihan.html', {
            'tagihan': tagihan,
            'madrasah_list': supervisor.madrasah_binaan.all(),
            'is_supervisor': True,
        })
    
    elif is_kepala_madrasah(request.user):
        try:
            madrasah = Madrasah.objects.get(kepala_madrasah__icontains=request.user.get_full_name())
            
            if request.method == 'POST' and 'upload_bukti' in request.FILES:
                tagihan_id = request.POST.get('tagihan_id')
                tagihan_obj = get_object_or_404(TagihanPengawas, id=tagihan_id, madrasah=madrasah)
                tagihan_obj.bukti_bayar = request.FILES['upload_bukti']
                tagihan_obj.status = 'dibayar'
                tagihan_obj.save()
                messages.success(request, 'Bukti pembayaran berhasil diupload!')
                return redirect('/kelola-tagihan/')
            
            tagihan = TagihanPengawas.objects.filter(madrasah=madrasah).order_by('-tanggal')
            return render(request, 'supervision/kelola_tagihan.html', {
                'tagihan': tagihan,
                'madrasah': madrasah,
                'is_supervisor': False,
            })
        except Madrasah.DoesNotExist:
            messages.warning(request, 'Data madrasah tidak ditemukan. Hubungi admin untuk menghubungkan akun Anda dengan data madrasah.')
            return redirect('/')  # ← Ganti dengan URL langsung
    
    return redirect('/')


# ============= LAPORAN PERIODE =============

def laporan_periode(request):
    from datetime import datetime
    
    # Ambil tahun dari data supervisi yang ada
    tahun_dari_supervisi = HasilSupervisi.objects.dates('tanggal', 'year').values_list('tanggal__year', flat=True)
    tahun_dari_manajerial = HasilSupervisiManajerial.objects.dates('tanggal', 'year').values_list('tanggal__year', flat=True)
    
    # Gabungkan dan ambil unik
    semua_tahun = set(list(tahun_dari_supervisi) + list(tahun_dari_manajerial))
    
    # Jika tidak ada data, gunakan tahun default
    if not semua_tahun:
        tahun_sekarang = datetime.now().year
        daftar_tahun = list(range(tahun_sekarang - 5, tahun_sekarang + 2))
    else:
        daftar_tahun = sorted(list(semua_tahun))
    
    return render(request, 'supervision/laporan_periode.html', {'daftar_tahun': daftar_tahun})


def export_laporan_pdf(request, jenis, tahun, periode=None):
    """Export laporan profesional lengkap multi-bab"""
    from datetime import datetime
    from django.db.models import Avg, Count, Q
    import json
    
    # Tentukan periode
    if jenis == 'bulanan' and periode:
        bulan_int = int(periode)
        nama_bulan = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
                      'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'][bulan_int-1]
        nama_periode = f"{nama_bulan} {tahun}"
        tanggal_start = datetime(int(tahun), bulan_int, 1)
        if bulan_int == 12:
            tanggal_end = datetime(int(tahun)+1, 1, 1)
        else:
            tanggal_end = datetime(int(tahun), bulan_int+1, 1)
    elif jenis == 'triwulan' and periode:
        triwulan_int = int(periode)
        if triwulan_int == 1:
            nama_periode = f"Triwulan I (Januari-Maret) {tahun}"
            tanggal_start = datetime(int(tahun), 1, 1)
            tanggal_end = datetime(int(tahun), 4, 1)
        elif triwulan_int == 2:
            nama_periode = f"Triwulan II (April-Juni) {tahun}"
            tanggal_start = datetime(int(tahun), 4, 1)
            tanggal_end = datetime(int(tahun), 7, 1)
        elif triwulan_int == 3:
            nama_periode = f"Triwulan III (Juli-September) {tahun}"
            tanggal_start = datetime(int(tahun), 7, 1)
            tanggal_end = datetime(int(tahun), 10, 1)
        else:
            nama_periode = f"Triwulan IV (Oktober-Desember) {tahun}"
            tanggal_start = datetime(int(tahun), 10, 1)
            tanggal_end = datetime(int(tahun)+1, 1, 1)
    elif jenis == 'semester' and periode:
        semester_int = int(periode)
        if semester_int == 1:
            nama_periode = f"Semester Ganjil {tahun}"
            tanggal_start = datetime(int(tahun), 1, 1)
            tanggal_end = datetime(int(tahun), 7, 1)
        else:
            nama_periode = f"Semester Genap {tahun}"
            tanggal_start = datetime(int(tahun), 7, 1)
            tanggal_end = datetime(int(tahun)+1, 1, 1)
    else:
        nama_periode = f"Tahun {tahun}"
        tanggal_start = datetime(int(tahun), 1, 1)
        tanggal_end = datetime(int(tahun)+1, 1, 1)
    
    # Ambil data supervisi dalam periode
    supervisi_akademik = HasilSupervisi.objects.filter(
        tanggal__gte=tanggal_start,
        tanggal__lt=tanggal_end
    )
    supervisi_manajerial = HasilSupervisiManajerial.objects.filter(
        tanggal__gte=tanggal_start,
        tanggal__lt=tanggal_end
    )
    
    # ============= STATISTIK UMUM =============
    total_madrasah = Madrasah.objects.count()
    total_guru = Guru.objects.count()
    total_supervisi_akademik = supervisi_akademik.count()
    total_supervisi_manajerial = supervisi_manajerial.count()
    
    # Rata-rata nilai
    rata_akademik = supervisi_akademik.aggregate(avg=Avg('nilai_total'))['avg'] or 0
    rata_manajerial = supervisi_manajerial.aggregate(avg=Avg('nilai_total'))['avg'] or 0
    
    # ============= DISTRIBUSI PREDIKAT =============
    predikat_akademik = {
        'A': supervisi_akademik.filter(nilai_total__gte=91).count(),
        'B': supervisi_akademik.filter(nilai_total__gte=76, nilai_total__lt=91).count(),
        'C': supervisi_akademik.filter(nilai_total__gte=61, nilai_total__lt=76).count(),
        'D': supervisi_akademik.filter(nilai_total__lt=61).count(),
    }
    
    predikat_manajerial = {
        'A': supervisi_manajerial.filter(nilai_total__gte=91).count(),
        'B': supervisi_manajerial.filter(nilai_total__gte=76, nilai_total__lt=91).count(),
        'C': supervisi_manajerial.filter(nilai_total__gte=61, nilai_total__lt=76).count(),
        'D': supervisi_manajerial.filter(nilai_total__lt=61).count(),
    }
    
    # ============= REKAP PER MADRASAH =============
    rekap_madrasah = []
    for m in Madrasah.objects.all():
        nilai_akademik = supervisi_akademik.filter(madrasah=m).aggregate(avg=Avg('nilai_total'))['avg'] or 0
        nilai_manajerial = supervisi_manajerial.filter(madrasah=m).aggregate(avg=Avg('nilai_total'))['avg'] or 0
        rekap_madrasah.append({
            'nama': m.nama,
            'npsn': m.npsn,
            'jenis': m.get_jenis_display(),
            'status': m.status,
            'kecamatan': m.kecamatan,
            'nilai_akademik': round(nilai_akademik, 2),
            'nilai_manajerial': round(nilai_manajerial, 2),
            'rata_rata': round((nilai_akademik + nilai_manajerial) / 2 if (nilai_akademik or nilai_manajerial) else 0, 2),
        })
    
    # Urutkan berdasarkan nilai
    rekap_madrasah.sort(key=lambda x: x['rata_rata'], reverse=True)
    
    # ============= REKAP PER KECAMATAN =============
    kecamatan_list = Madrasah.objects.values_list('kecamatan', flat=True).distinct()
    rekap_kecamatan = []
    for kec in kecamatan_list:
        if kec:
            madrasah_kec = Madrasah.objects.filter(kecamatan=kec)
            nilai_akademik = supervisi_akademik.filter(madrasah__in=madrasah_kec).aggregate(avg=Avg('nilai_total'))['avg'] or 0
            nilai_manajerial = supervisi_manajerial.filter(madrasah__in=madrasah_kec).aggregate(avg=Avg('nilai_total'))['avg'] or 0
            rekap_kecamatan.append({
                'nama': kec,
                'jumlah_madrasah': madrasah_kec.count(),
                'nilai_akademik': round(nilai_akademik, 2),
                'nilai_manajerial': round(nilai_manajerial, 2),
                'rata_rata': round((nilai_akademik + nilai_manajerial) / 2 if (nilai_akademik or nilai_manajerial) else 0, 2),
            })
    rekap_kecamatan.sort(key=lambda x: x['rata_rata'], reverse=True)
    
    # ============= GRAFIK DATA (JSON) =============
    # Data per bulan
    bulan_data = []
    for i in range(1, 13):
        count = supervisi_akademik.filter(tanggal__month=i).count()
        bulan_data.append(count)
    
    # Data per jenis madrasah
    jenis_data = []
    for jenis_code, jenis_name in [('RA', 'RA'), ('MI', 'MI'), ('MTs', 'MTs'), ('MA', 'MA')]:
        count = Madrasah.objects.filter(jenis=jenis_code).count()
        jenis_data.append({'jenis': jenis_name, 'jumlah': count})
    
    # ============= BUAT HTML LAPORAN =============
    html = f"""
<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <title>Laporan Pengawasan Madrasah - {nama_periode}</title>
    <style>
        @page {{
            size: A4;
            margin: 2.5cm;
        }}
        body {{
            font-family: 'Times New Roman', 'Calibri', sans-serif;
            line-height: 1.5;
            color: #333;
            font-size: 12pt;
            text-align: justify;
        }}
        h1 {{ text-align: center; font-size: 18pt; margin-bottom: 20px; }}
        h2 {{ font-size: 16pt; margin-top: 25px; margin-bottom: 15px; border-bottom: 2px solid #2c3e50; padding-bottom: 5px; }}
        h3 {{ font-size: 14pt; margin-top: 20px; margin-bottom: 10px; }}
        h4 {{ font-size: 12pt; margin-top: 15px; margin-bottom: 8px; }}
        .cover {{
            text-align: center;
            margin-top: 100px;
            margin-bottom: 100px;
        }}
        .cover h1 {{ font-size: 24pt; margin-bottom: 40px; }}
        .cover h2 {{ font-size: 18pt; border: none; margin-bottom: 30px; }}
        .cover .title {{ margin-top: 80px; }}
        .cover .bottom {{ margin-top: 80px; }}
        .lembar-pengesahan {{
            margin-top: 80px;
            text-align: center;
        }}
        .signature {{
            margin-top: 50px;
            text-align: right;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }}
        th {{
            background-color: #2c3e50;
            color: white;
        }}
        .table-bordered th, .table-bordered td {{ border: 1px solid #ddd; }}
        .text-center {{ text-align: center; }}
        .text-right {{ text-align: right; }}
        .footer {{
            text-align: center;
            margin-top: 50px;
            font-size: 10pt;
            border-top: 1px solid #ddd;
            padding-top: 20px;
        }}
        .page-break {{
            page-break-before: always;
        }}
        .badge-success {{ background-color: #28a745; color: white; padding: 2px 8px; border-radius: 4px; }}
        .badge-primary {{ background-color: #007bff; color: white; padding: 2px 8px; border-radius: 4px; }}
        .badge-warning {{ background-color: #ffc107; color: #333; padding: 2px 8px; border-radius: 4px; }}
        .badge-danger {{ background-color: #dc3545; color: white; padding: 2px 8px; border-radius: 4px; }}
        .ringkasan {{
            background-color: #f8f9fa;
            padding: 15px;
            margin: 20px 0;
            border-left: 4px solid #2c3e50;
        }}
        .temuan {{
            background-color: #fff3cd;
            padding: 15px;
            margin: 15px 0;
            border-left: 4px solid #ffc107;
        }}
    </style>
</head>
<body>

<!-- ==================== COVER ==================== -->
<div class="cover">
    <h1>LAPORAN PENGAWASAN MADRASAH</h1>
    <h2>{nama_periode}</h2>
    <div class="title">
        <p>Disusun oleh:</p>
        <p><strong>Pengawas Madrasah</strong></p>
        <p>Dinas Pendidikan Kabupaten/Kota</p>
    </div>
    <div class="bottom">
        <p>{datetime.now().strftime('%Y')}</p>
    </div>
</div>

<div class="page-break"></div>

<!-- ==================== LEMBAR PENGESAHAN ==================== -->
<div class="lembar-pengesahan">
    <h1>LEMBAR PENGESAHAN</h1>
    <p style="margin-top: 50px;">Laporan Pengawasan Madrasah periode <strong>{nama_periode}</strong> telah disusun dan disahkan.</p>
    <div class="signature">
        <p>Mengetahui,</p>
        <p>Kepala Dinas Pendidikan Kabupaten/Kota</p>
        <br><br><br>
        <p><u>(Nama Kepala Dinas)</u></p>
        <p>NIP. xxxxxxxxxxxxxxx</p>
    </div>
</div>

<div class="page-break"></div>

<!-- ==================== KATA PENGANTAR ==================== -->
<h1>KATA PENGANTAR</h1>
<p>Puji syukur kita panjatkan ke hadirat Allah SWT, Tuhan Yang Maha Esa, karena atas rahmat dan karunia-Nya, laporan pengawasan madrasah periode <strong>{nama_periode}</strong> ini dapat diselesaikan dengan baik.</p>
<p>Laporan ini disusun sebagai bentuk pertanggungjawaban pelaksanaan tugas pengawasan madrasah di Kabupaten/Kota. Pengawasan madrasah merupakan upaya sistematis untuk menjamin mutu pendidikan madrasah melalui kegiatan supervisi akademik dan manajerial.</p>
<p>Kami mengucapkan terima kasih kepada semua pihak yang telah membantu dalam pelaksanaan pengawasan dan penyusunan laporan ini, terutama kepada Kepala Dinas Pendidikan, para Kepala Madrasah, guru, serta seluruh pemangku kepentingan yang telah berkolaborasi dengan baik.</p>
<p>Kami menyadari bahwa laporan ini masih jauh dari sempurna. Oleh karena itu, kritik dan saran yang membangun sangat kami harapkan untuk perbaikan di masa mendatang.</p>
<p>Semoga laporan ini bermanfaat bagi peningkatan mutu pendidikan madrasah di Kabupaten/Kota.</p>
<div class="signature">
    <p>Kabupaten/Kota, {datetime.now().strftime('%d %B %Y')}</p>
    <p>Tim Pengawas Madrasah</p>
</div>

<div class="page-break"></div>

<!-- ==================== DAFTAR ISI ==================== -->
<h1>DAFTAR ISI</h1>
<ul>
    <li>Halaman Judul ............................................. i</li>
    <li>Lembar Pengesahan ........................................ ii</li>
    <li>Kata Pengantar ........................................... iii</li>
    <li>Daftar Isi ............................................... iv</li>
    <li>Daftar Tabel ............................................. v</li>
    <li>Daftar Grafik ............................................ vi</li>
    <li><strong>BAB I PENDAHULUAN</strong> ........................ 1</li>
    <li><strong>BAB II GAMBARAN UMUM</strong> ..................... 5</li>
    <li><strong>BAB III HASIL SUPERVISI</strong> .................. 9</li>
    <li><strong>BAB IV PEMBAHASAN DAN TEMUAN</strong> .............. 15</li>
    <li><strong>BAB V KESIMPULAN DAN REKOMENDASI</strong> .......... 20</li>
    <li>Daftar Pustaka ........................................... 24</li>
    <li>Lampiran ................................................. 25</li>
</ul>

<div class="page-break"></div>

<!-- ==================== BAB I - PENDAHULUAN ==================== -->
<h1>BAB I</h1>
<h1>PENDAHULUAN</h1>

<h2>1.1 Latar Belakang</h2>
<p>Madrasah sebagai lembaga pendidikan berciri khas Islam memiliki peran strategis dalam mencerdaskan kehidupan bangsa dan membentuk karakter peserta didik yang beriman, bertakwa, dan berakhlak mulia. Untuk menjamin mutu pendidikan madrasah, diperlukan pengawasan yang sistematis, terencana, dan berkelanjutan.</p>
<p>Berdasarkan Peraturan Menteri Agama (PMA) Nomor 2 Tahun 2012 tentang Pengawas Madrasah dan Peraturan Menteri Pendayagunaan Aparatur Negara dan Reformasi Birokrasi (PermenPAN RB) Nomor 21 Tahun 2024, pengawas madrasah memiliki tugas melaksanakan supervisi akademik dan supervisi manajerial secara berkala.</p>
<p>Periode <strong>{nama_periode}</strong>, telah dilaksanakan kegiatan pengawasan terhadap <strong>{total_madrasah} madrasah</strong> yang tersebar di Kabupaten/Kota. Laporan ini disusun untuk mendokumentasikan hasil pengawasan sebagai bahan evaluasi dan perbaikan mutu.</p>

<h2>1.2 Tujuan</h2>
<p>Adapun tujuan pelaksanaan pengawasan madrasah pada periode ini adalah:</p>
<ol>
    <li>Menilai kinerja guru dalam melaksanakan pembelajaran (supervisi akademik);</li>
    <li>Menilai kinerja kepala madrasah dalam mengelola lembaga (supervisi manajerial);</li>
    <li>Mengidentifikasi kelemahan dan kekuatan pengelolaan madrasah;</li>
    <li>Memberikan rekomendasi perbaikan untuk peningkatan mutu madrasah;</li>
    <li>Menyusun program tindak lanjut pengawasan.</li>
</ol>

<h2>1.3 Ruang Lingkup</h2>
<p>Ruang lingkup pengawasan pada periode ini meliputi:</p>
<ol>
    <li><strong>Supervisi Akademik</strong> terhadap <strong>{total_guru} guru</strong> dengan fokus pada perencanaan, pelaksanaan, penilaian, dan tindak lanjut pembelajaran;</li>
    <li><strong>Supervisi Manajerial</strong> terhadap <strong>{total_madrasah} madrasah</strong> dengan fokus pada 8 Standar Nasional Pendidikan (SNP).</li>
</ol>

<h2>1.4 Metodologi</h2>
<p>Metode yang digunakan dalam pelaksanaan pengawasan meliputi:</p>
<ol>
    <li>Observasi langsung ke madrasah;</li>
    <li>Wawancara dengan kepala madrasah, guru, dan komite madrasah;</li>
    <li>Studi dokumentasi (RPP, Kurikulum, RKAM, dll);</li>
    <li>Kuesioner/angket penilaian;</li>
    <li>Analisis data menggunakan Sistem Pengawasan Madrasah Online.</li>
</ol>

<div class="page-break"></div>

<!-- ==================== BAB II - GAMBARAN UMUM ==================== -->
<h1>BAB II</h1>
<h1>GAMBARAN UMUM</h1>

<h2>2.1 Profil Madrasah di Kabupaten/Kota</h2>
<p>Kabupaten/Kota memiliki <strong>{total_madrasah} madrasah</strong> yang tersebar di berbagai kecamatan. Berikut adalah sebaran madrasah berdasarkan jenjang dan status:</p>

<h3>Tabel 2.1 Sebaran Madrasah Berdasarkan Jenjang</h3>
<table class="table-bordered">
    <thead>
        <tr><th>Jenjang</th><th>Jumlah</th><th>Persentase</th></tr>
    </thead>
    <tbody>
        {''.join([f"<tr><td>{d['jenis']}</td><td>{d['jumlah']}</td><td>{round(d['jumlah']/total_madrasah*100, 1) if total_madrasah > 0 else 0}%</td></tr>" for d in jenis_data])}
        <tr style="background-color:#f2f2f2;"><td><strong>Total</strong></td><td><strong>{total_madrasah}</strong></td><td><strong>100%</strong></td></tr>
    </tbody>
</table>

<h3>Tabel 2.2 Sebaran Madrasah Berdasarkan Status</h3>
<table class="table-bordered">
    <thead>
        <tr><th>Status</th><th>Jumlah</th><th>Persentase</th></tr>
    </thead>
    <tbody>
        <tr><td>Negeri</td><td>{Madrasah.objects.filter(status='Negeri').count()}</td><td>{round(Madrasah.objects.filter(status='Negeri').count()/total_madrasah*100, 1) if total_madrasah > 0 else 0}%</td></tr>
        <tr><td>Swasta</td><td>{Madrasah.objects.filter(status='Swasta').count()}</td><td>{round(Madrasah.objects.filter(status='Swasta').count()/total_madrasah*100, 1) if total_madrasah > 0 else 0}%</td></tr>
    </tbody>
</table>

<h2>2.2 Profil Guru</h2>
<p>Jumlah guru yang terdaftar pada periode ini sebanyak <strong>{total_guru} orang</strong>.</p>

<div class="page-break"></div>

<!-- ==================== BAB III - HASIL SUPERVISI ==================== -->
<h1>BAB III</h1>
<h1>HASIL SUPERVISI</h1>

<h2>3.1 Supervisi Akademik</h2>
<p>Selama periode {nama_periode}, telah dilaksanakan <strong>{total_supervisi_akademik} kali supervisi akademik</strong> terhadap guru. Berikut ringkasan hasilnya:</p>

<h3>Tabel 3.1 Ringkasan Nilai Supervisi Akademik</h3>
<table class="table-bordered">
    <thead><tr><th>Indikator</th><th>Jumlah</th><th>Persentase</th></tr></thead>
    <tbody>
        <tr><td>Total Supervisi</td><td>{total_supervisi_akademik}</td><td>100%</td></tr>
        <tr><td>Rata-rata Nilai</td><td colspan="2">{rata_akademik:.2f}</td></tr>
    </tbody>
</table>

<h3>Tabel 3.2 Distribusi Predikat Supervisi Akademik</h3>
<table class="table-bordered">
    <thead><tr><th>Predikat</th><th>Kisaran Nilai</th><th>Jumlah</th><th>Persentase</th></tr></thead>
    <tbody>
        <tr><td><span class="badge-success">A (Sangat Baik)</span></td><td>91-100</td><td>{predikat_akademik['A']}</td><td>{round(predikat_akademik['A']/total_supervisi_akademik*100, 1) if total_supervisi_akademik > 0 else 0}%</td></tr>
        <tr><td><span class="badge-primary">B (Baik)</span></td><td>76-90</td><td>{predikat_akademik['B']}</td><td>{round(predikat_akademik['B']/total_supervisi_akademik*100, 1) if total_supervisi_akademik > 0 else 0}%</td></tr>
        <tr><td><span class="badge-warning">C (Cukup)</span></td><td>61-75</td><td>{predikat_akademik['C']}</td><td>{round(predikat_akademik['C']/total_supervisi_akademik*100, 1) if total_supervisi_akademik > 0 else 0}%</td></tr>
        <tr><td><span class="badge-danger">D (Kurang)</span></td><td>≤60</td><td>{predikat_akademik['D']}</td><td>{round(predikat_akademik['D']/total_supervisi_akademik*100, 1) if total_supervisi_akademik > 0 else 0}%</td></tr>
    </tbody>
</table>

<h2>3.2 Supervisi Manajerial</h2>
<p>Selama periode {nama_periode}, telah dilaksanakan <strong>{total_supervisi_manajerial} kali supervisi manajerial</strong> terhadap kepala madrasah. Berikut ringkasan hasilnya:</p>

<h3>Tabel 3.3 Ringkasan Nilai Supervisi Manajerial</h3>
<table class="table-bordered">
    <thead><tr><th>Indikator</th><th>Jumlah</th><th>Persentase</th></tr></thead>
    <tbody>
        <tr><td>Total Supervisi</td><td>{total_supervisi_manajerial}</td><td>100%</td></tr>
        <tr><td>Rata-rata Nilai</td><td colspan="2">{rata_manajerial:.2f}</td></tr>
    </tbody>
</table>

<h3>Tabel 3.4 Distribusi Predikat Supervisi Manajerial</h3>
<table class="table-bordered">
    <thead><tr><th>Predikat</th><th>Kisaran Nilai</th><th>Jumlah</th><th>Persentase</th></tr></thead>
    <tbody>
        <tr><td><span class="badge-success">A (Sangat Baik)</span></td><td>91-100</td><td>{predikat_manajerial['A']}</td><td>{round(predikat_manajerial['A']/total_supervisi_manajerial*100, 1) if total_supervisi_manajerial > 0 else 0}%</td></tr>
        <tr><td><span class="badge-primary">B (Baik)</span></td><td>76-90</td><td>{predikat_manajerial['B']}</td><td>{round(predikat_manajerial['B']/total_supervisi_manajerial*100, 1) if total_supervisi_manajerial > 0 else 0}%</td></tr>
        <tr><td><span class="badge-warning">C (Cukup)</span></td><td>61-75</td><td>{predikat_manajerial['C']}</td><td>{round(predikat_manajerial['C']/total_supervisi_manajerial*100, 1) if total_supervisi_manajerial > 0 else 0}%</td></tr>
        <tr><td><span class="badge-danger">D (Kurang)</span></td><td>≤60</td><td>{predikat_manajerial['D']}</td><td>{round(predikat_manajerial['D']/total_supervisi_manajerial*100, 1) if total_supervisi_manajerial > 0 else 0}%</td></tr>
    </tbody>
</table>

<h2>3.3 Rekap Nilai per Madrasah</h2>
<table class="table-bordered" style="font-size: 10pt;">
    <thead>
        <tr><th>No</th><th>NPSN</th><th>Nama Madrasah</th><th>Kecamatan</th><th>Nilai Akademik</th><th>Nilai Manajerial</th><th>Rata-rata</th><th>Status</th></tr>
    </thead>
    <tbody>
        {''.join([f"<tr><td>{i+1}</td><td>{m['npsn']}</td><td>{m['nama']}</td><td>{m['kecamatan']}</td><td>{m['nilai_akademik']}</td><td>{m['nilai_manajerial']}</td><td><strong>{m['rata_rata']}</strong></td><td>{'🟢' if m['rata_rata'] >= 76 else '🟡' if m['rata_rata'] >= 61 else '🔴'}</td></tr>" for i, m in enumerate(rekap_madrasah[:20])])}
    </tbody>
</table>
<p><small>Catatan: Ditampilkan 20 madrasah teratas. Data lengkap tersedia di lampiran.</small></p>

<h2>3.4 Rekap Nilai per Kecamatan</h2>
<table class="table-bordered">
    <thead>
        <tr><th>No</th><th>Kecamatan</th><th>Jumlah Madrasah</th><th>Nilai Akademik</th><th>Nilai Manajerial</th><th>Rata-rata</th><th>Kategori</th></tr>
    </thead>
    <tbody>
        {''.join([f"<tr><td>{i+1}</td><td>{k['nama']}</td><td>{k['jumlah_madrasah']}</td><td>{k['nilai_akademik']}</td><td>{k['nilai_manajerial']}</td><td><strong>{k['rata_rata']}</strong></td><td>{'Sangat Baik' if k['rata_rata'] >= 91 else 'Baik' if k['rata_rata'] >= 76 else 'Cukup' if k['rata_rata'] >= 61 else 'Kurang'}</td></tr>" for i, k in enumerate(rekap_kecamatan)])}
    </tbody>
</table>

<div class="page-break"></div>

<!-- ==================== BAB IV - PEMBAHASAN ==================== -->
<h1>BAB IV</h1>
<h1>PEMBAHASAN DAN TEMUAN</h1>

<h2>4.1 Analisis Hasil Supervisi Akademik</h2>
<p>Berdasarkan hasil supervisi akademik, secara umum kinerja guru di Kabupaten/Kota berada pada kategori <strong>{'Sangat Baik' if rata_akademik >= 91 else 'Baik' if rata_akademik >= 76 else 'Cukup' if rata_akademik >= 61 else 'Kurang'}</strong> dengan rata-rata nilai <strong>{rata_akademik:.2f}</strong>.</p>

<div class="ringkasan">
    <h4>Ringkasan Temuan Supervisi Akademik:</h4>
    <ul>
        <li>Sebanyak <strong>{predikat_akademik['A']} supervisi ({round(predikat_akademik['A']/total_supervisi_akademik*100, 1) if total_supervisi_akademik > 0 else 0}%)</strong> mencapai predikat A (Sangat Baik)</li>
        <li>Sebanyak <strong>{predikat_akademik['D']} supervisi ({round(predikat_akademik['D']/total_supervisi_akademik*100, 1) if total_supervisi_akademik > 0 else 0}%)</strong> masih pada predikat D (Kurang) - perlu pendampingan intensif</li>
        <li>Komponen terkuat: <strong>Perencanaan Pembelajaran</strong> dengan rata-rata nilai tertinggi</li>
        <li>Komponen terlemah: <strong>Tindak Lanjut</strong> - masih banyak guru yang belum melaksanakan remedial/pengayaan secara terprogram</li>
    </ul>
</div>

<h2>4.2 Analisis Hasil Supervisi Manajerial</h2>
<p>Berdasarkan hasil supervisi manajerial, pengelolaan madrasah di Kabupaten/Kota berada pada kategori <strong>{'Sangat Baik' if rata_manajerial >= 91 else 'Baik' if rata_manajerial >= 76 else 'Cukup' if rata_manajerial >= 61 else 'Kurang'}</strong> dengan rata-rata nilai <strong>{rata_manajerial:.2f}</strong>.</p>

<div class="ringkasan">
    <h4>Ringkasan Temuan Supervisi Manajerial:</h4>
    <ul>
        <li>Sebanyak <strong>{predikat_manajerial['A']} supervisi ({round(predikat_manajerial['A']/total_supervisi_manajerial*100, 1) if total_supervisi_manajerial > 0 else 0}%)</strong> mencapai predikat A (Sangat Baik)</li>
        <li>Standar yang paling baik terpenuhi: <strong>Standar Isi dan Standar Proses</strong></li>
        <li>Standar yang masih lemah: <strong>Standar Sarana Prasarana dan Standar Pembiayaan</strong></li>
        <li>Masih terdapat madrasah yang belum memiliki dokumen RKAM yang lengkap</li>
    </ul>
</div>

<h2>4.3 Temuan Khusus</h2>
<div class="temuan">
    <h4>🔍 Temuan Penting:</h4>
    <ol>
        <li><strong>Kesenjangan Mutu Antar Kecamatan</strong> - Terdapat perbedaan signifikan antara kecamatan dengan nilai tertinggi ({rekap_kecamatan[0]['nama'] if rekap_kecamatan else '-'}: {rekap_kecamatan[0]['rata_rata'] if rekap_kecamatan else 0}) dan terendah ({rekap_kecamatan[-1]['nama'] if rekap_kecamatan else '-'}: {rekap_kecamatan[-1]['rata_rata'] if rekap_kecamatan else 0})</li>
        <li><strong>Keterbatasan Sarana Prasarana</strong> - Beberapa madrasah masih kekurangan laboratorium dan perpustakaan yang memadai</li>
        <li><strong>Kompetensi Guru dalam HOTS</strong> - Masih rendahnya kemampuan guru dalam mengembangkan soal Higher Order Thinking Skills (HOTS)</li>
    </ol>
</div>

<div class="page-break"></div>

<!-- ==================== BAB V - KESIMPULAN ==================== -->
<h1>BAB V</h1>
<h1>KESIMPULAN DAN REKOMENDASI</h1>

<h2>5.1 Kesimpulan</h2>
<p>Berdasarkan hasil pengawasan madrasah pada periode <strong>{nama_periode}</strong>, dapat disimpulkan sebagai berikut:</p>
<ol>
    <li>Secara umum, kinerja guru dalam melaksanakan pembelajaran berada pada kategori <strong>{'Sangat Baik' if rata_akademik >= 91 else 'Baik' if rata_akademik >= 76 else 'Cukup' if rata_akademik >= 61 else 'Kurang'}</strong> dengan rata-rata nilai <strong>{rata_akademik:.2f}</strong>.</li>
    <li>Secara umum, pengelolaan madrasah berada pada kategori <strong>{'Sangat Baik' if rata_manajerial >= 91 else 'Baik' if rata_manajerial >= 76 else 'Cukup' if rata_manajerial >= 61 else 'Kurang'}</strong> dengan rata-rata nilai <strong>{rata_manajerial:.2f}</strong>.</li>
    <li>Masih terdapat kesenjangan mutu antar kecamatan yang perlu mendapatkan perhatian khusus.</li>
    <li>Supervisi akademik dan manajerial telah dilaksanakan sesuai dengan jadwal dan prosedur yang ditetapkan.</li>
</ol>

<h2>5.2 Rekomendasi</h2>
<p>Berdasarkan temuan di atas, berikut rekomendasi yang dapat ditindaklanjuti:</p>

<h3>5.2.1 Untuk Guru</h3>
<ol>
    <li>Meningkatkan kemampuan dalam menyusun perangkat pembelajaran yang berorientasi HOTS;</li>
    <li>Melaksanakan program remedial dan pengayaan secara terprogram;</li>
    <li>Mengikuti pelatihan dan pengembangan keprofesian berkelanjutan (PKB);</li>
    <li>Memanfaatkan teknologi dalam pembelajaran (TPACK).</li>
</ol>

<h3>5.2.2 Untuk Kepala Madrasah</h3>
<ol>
    <li>Melengkapi dokumen administrasi madrasah (KOM, RKAM, dll);</li>
    <li>Meningkatkan pemenuhan Standar Sarana Prasarana secara bertahap;</li>
    <li>Mengoptimalkan peran komite madrasah dalam pengelolaan;</li>
    <li>Melaksanakan evaluasi diri madrasah (EDM) secara berkala.</li>
</ol>

<h3>5.2.3 Untuk Dinas Pendidikan</h3>
<ol>
    <li>Memprioritaskan pendampingan pada madrasah dengan nilai di bawah 70;</li>
    <li>Memfasilitasi pelatihan peningkatan kompetensi guru dan kepala madrasah;</li>
    <li>Mengalokasikan anggaran untuk peningkatan sarana prasarana madrasah;</li>
    <li>Melaksanakan monitoring dan evaluasi secara berkelanjutan.</li>
</ol>

<h2>5.3 Rencana Tindak Lanjut</h2>
<table class="table-bordered">
    <thead>
        <tr><th>No</th><th>Kegiatan</th><th>Target</th><th>Waktu</th><th>Penanggung Jawab</th></tr>
    </thead>
    <tbody>
        <tr><td>1</td><td>Pendampingan madrasah nilai rendah</td><td>Semua madrasah nilai <70</td><td>1 bulan</td><td>Pengawas</td></tr>
        <tr><td>2</td><td>Pelatihan penyusunan perangkat HOTS</td><td>Semua guru</td><td>2 bulan</td><td>Dinas Pendidikan</td></tr>
        <tr><td>3</td><td>Evaluasi tengah semester</td><td>Semua madrasah</td><td>3 bulan</td><td>Pengawas & Kepsek</td></tr>
        <tr><td>4</td><td>Monitoring pemenuhan 8 SNP</td><td>Semua madrasah</td><td>6 bulan</td><td>Dinas Pendidikan</td></tr>
    </tbody>
</table>

<div class="page-break"></div>

<!-- ==================== DAFTAR PUSTAKA ==================== -->
<h1>DAFTAR PUSTAKA</h1>
<ul>
    <li>Peraturan Menteri Agama Republik Indonesia Nomor 2 Tahun 2012 tentang Pengawas Madrasah.</li>
    <li>Peraturan Menteri Pendayagunaan Aparatur Negara dan Reformasi Birokrasi Nomor 21 Tahun 2024 tentang Jabatan Fungsional Pengawas Madrasah.</li>
    <li>Peraturan Pemerintah Nomor 19 Tahun 2005 tentang Standar Nasional Pendidikan.</li>
    <li>Peraturan Menteri Pendidikan Nasional Nomor 12 Tahun 2007 tentang Standar Pengawas Sekolah/Madrasah.</li>
    <li>Undang-Undang Republik Indonesia Nomor 20 Tahun 2003 tentang Sistem Pendidikan Nasional.</li>
</ul>

<div class="page-break"></div>

<!-- ==================== LAMPIRAN ==================== -->
<h1>LAMPIRAN</h1>

<h2>Lampiran 1: Rekap Lengkap Nilai per Madrasah</h2>
<table class="table-bordered" style="font-size: 9pt;">
    <thead>
        <tr><th>No</th><th>NPSN</th><th>Nama Madrasah</th><th>Kecamatan</th><th>Nilai Akademik</th><th>Nilai Manajerial</th><th>Rata-rata</th></tr>
    </thead>
    <tbody>
        {''.join([f"<tr><td>{i+1}</td><td>{m['npsn']}</td><td>{m['nama']}</td><td>{m['kecamatan']}</td><td>{m['nilai_akademik']}</td><td>{m['nilai_manajerial']}</td><td><strong>{m['rata_rata']}</strong></td></table>" for i, m in enumerate(rekap_madrasah)])}
    </tbody>
</table>

<h2>Lampiran 2: Instrumen Supervisi yang Digunakan</h2>
<p>Instrumen supervisi yang digunakan mengacu pada regulasi terbaru Kemenag RI, meliputi:</p>
<ul>
    <li>Instrumen Supervisi Akademik (untuk guru) - {InstrumenSupervisi.objects.count()} indikator</li>
    <li>Instrumen Supervisi Manajerial (untuk kepala madrasah) - {InstrumenManajerial.objects.count()} indikator</li>
</ul>

<div class="footer">
    <p>Laporan ini dihasilkan secara otomatis dari Sistem Pengawasan Madrasah Online</p>
    <p>{datetime.now().strftime('%d %B %Y')}</p>
</div>

</body>
</html>
    """
    
    response = HttpResponse(html, content_type='text/html')
    response['Content-Disposition'] = f'inline; filename="Laporan_Pengawasan_Madrasah_{nama_periode.replace(" ", "_")}.html"'
    return response

def export_laporan_tahunan(request, tahun):
    """Export laporan tahunan"""
    return export_laporan_pdf(request, 'tahunan', tahun)

# ============= KEPALA MADRASAH =============

@login_required
@user_passes_test(is_kepala_madrasah)
def hasil_supervisi_kepsek(request):
    nama_kepala = request.user.get_full_name()
    
    try:
        madrasah = Madrasah.objects.get(kepala_madrasah__icontains=nama_kepala)
    except Madrasah.DoesNotExist:
        messages.warning(request, 'Data madrasah tidak ditemukan')
        return redirect('/')
    
    supervisi_akademik = HasilSupervisi.objects.filter(madrasah=madrasah).order_by('-tanggal')
    supervisi_manajerial = HasilSupervisiManajerial.objects.filter(madrasah=madrasah).order_by('-tanggal')
    
    return render(request, 'supervision/hasil_supervisi_kepsek.html', {
        'madrasah': madrasah,
        'supervisi_akademik': supervisi_akademik,
        'supervisi_manajerial': supervisi_manajerial,
    })


@login_required
@user_passes_test(is_kepala_madrasah)
def detail_supervisi_kepsek(request, pk, jenis):
    nama_kepala = request.user.get_full_name()
    
    try:
        madrasah = Madrasah.objects.get(kepala_madrasah__icontains=nama_kepala)
    except Madrasah.DoesNotExist:
        return redirect('/')
    
    if jenis == 'akademik':
        supervisi = get_object_or_404(HasilSupervisi, pk=pk, madrasah=madrasah)
    else:
        supervisi = get_object_or_404(HasilSupervisiManajerial, pk=pk, madrasah=madrasah)
    
    return render(request, 'supervision/detail_supervisi_kepsek.html', {'supervisi': supervisi, 'madrasah': madrasah})


# ============= REKAP KABUPATEN =============

@staff_member_required
def dashboard_rekap_kabupaten(request):
    from django.db.models import Avg
    
    context = {
        'total_madrasah': Madrasah.objects.count(),
        'total_guru': Guru.objects.count(),
        'total_supervisi': HasilSupervisi.objects.count(),
        'rata_akademik': round(HasilSupervisi.objects.aggregate(Avg('nilai_total'))['nilai_total__avg'] or 0, 2),
    }
    return render(request, 'supervision/dashboard_rekap_kabupaten.html', context)


@staff_member_required
def export_lengkap_excel(request):
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Madrasah"
    
    ws['A1'] = 'No'
    ws['B1'] = 'NPSN'
    ws['C1'] = 'Nama Madrasah'
    ws['D1'] = 'Kecamatan'
    ws['E1'] = 'Jenis'
    ws['F1'] = 'Status'
    
    for idx, m in enumerate(Madrasah.objects.all(), 2):
        ws[f'A{idx}'] = idx-1
        ws[f'B{idx}'] = m.npsn
        ws[f'C{idx}'] = m.nama
        ws[f'D{idx}'] = m.kecamatan
        ws[f'E{idx}'] = m.get_jenis_display()
        ws[f'F{idx}'] = m.status
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Rekap_Pengawasan_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def kirim_notifikasi_email(request):
    from django.core.mail import send_mail
    from django.conf import settings
    
    try:
        send_mail(
            'Laporan Pengawasan Madrasah',
            'Laporan pengawasan telah tersedia. Silakan login ke sistem.',
            settings.EMAIL_HOST_USER,
            [settings.EMAIL_HOST_USER],
            fail_silently=False,
        )
        messages.success(request, 'Notifikasi email berhasil dikirim!')
    except Exception as e:
        messages.error(request, f'Gagal mengirim email: {str(e)}')
    
    return redirect('dashboard_rekap_kabupaten')